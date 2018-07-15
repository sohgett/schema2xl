# -*- coding: utf-8 -*-

import openpyxl
import pymysql

DB_HOST = 'localhost'
DB_PORT = 3306
DB_USER = 'user'
DB_PASSWD = 'pass'
DB_NAME = 'test'
DB_CHARSET = 'utf8'
XLSX_NAME = 'schema2xl.xlsx'

CUSTOME_FILTER = ''

conn = None


def dbconnect(host=DB_HOST, port=DB_PORT, user=DB_USER,
              passwd=DB_PASSWD, db=DB_NAME, charset=DB_CHARSET):
    global conn
    if conn:
        return
    conn = pymysql.connect(host=host, port=port, user=user,
                           passwd=passwd, db=db, charset=charset)


def dbcursor():
    return conn.cursor(pymysql.cursors.DictCursor)


def fetch_columns(db=DB_NAME, filter=CUSTOME_FILTER):
    cur = dbcursor()
    s = []
    s.append("select")
    s.append("C.*, T.TABLE_COMMENT, K.REFERENCED_TABLE_NAME")
    s.append("from information_schema.COLUMNS as C")
    s.append("inner join information_schema.TABLES as T")
    s.append("on T.TABLE_NAME = C.TABLE_NAME")
    s.append("left outer join information_schema.KEY_COLUMN_USAGE as K")
    s.append("on K.TABLE_NAME = C.TABLE_NAME")
    s.append("and K.COLUMN_NAME = C.COLUMN_NAME")
    s.append("and K.POSITION_IN_UNIQUE_CONSTRAINT is not NULL")
    s.append("and K.TABLE_SCHEMA = %(db)s")
    s.append("where T.TABLE_TYPE = 'BASE TABLE'")
    s.append("and C.TABLE_SCHEMA = %(db)s")
    s.append("and T.TABLE_SCHEMA = %(db)s")
    s.append(filter)
    s.append("order by C.TABLE_NAME, C.ORDINAL_POSITION")
    cur.execute('\n'.join(s), dict(db=db))
    tables = []
    for row in cur:
        tables.append(row)
    cur.close()
    return tables


def write_xlsx(columns, xlsx=XLSX_NAME):
    def style_caption(c):
        c.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='top',
        )
        c.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin', color='FF000000'),
            right=openpyxl.styles.Side(border_style='thin', color='FF000000'),
            top=openpyxl.styles.Side(border_style='thin', color='FF000000'),
            bottom=openpyxl.styles.Side(border_style='thin', color='FF000000'),
        )
        c.fill = openpyxl.styles.PatternFill(
            fill_type='solid',
            start_color='FFE6E6E6',
            end_color='FFE6E6E6',
        )
        c.font = openpyxl.styles.Font(
            bold=True,
            name=u'ＭＳ Ｐゴシック',
        )

    def draw_caption():
        ws['A1'].value = u'テーブル名'
        ws['B1'].value = u'テーブルコメント'
        ws['C1'].value = u'SEQ'
        ws['D1'].value = u'カラム名'
        ws['E1'].value = u'PK'
        ws['F1'].value = u'FK'
        ws['G1'].value = u'必須'
        ws['H1'].value = u'カラムコメント'
        ws['I1'].value = u'データ型'
        for row in ws['A1:I1']:
            for c in row:
                style_caption(c)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 38
        ws.column_dimensions['E'].width = 5
        ws.column_dimensions['F'].width = 5
        ws.column_dimensions['G'].width = 5
        ws.column_dimensions['H'].width = 32
        ws.column_dimensions['I'].width = 16

    def style_column(c, wrap_text, horizontal):
        c.alignment = openpyxl.styles.Alignment(
            wrap_text=wrap_text,
            vertical='top',
            horizontal=horizontal,
        )
        c.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            top=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin'),
        )
        c.font = openpyxl.styles.Font(
            name=u'ＭＳ Ｐゴシック',
        )

    def draw_column(column, n):
        ws['A{0}'.format(n)].value = column['TABLE_NAME']
        ws['B{0}'.format(n)].value = column['TABLE_COMMENT']
        ws['C{0}'.format(n)].value = column['ORDINAL_POSITION']
        ws['D{0}'.format(n)].value = column['COLUMN_NAME']
        if column['COLUMN_KEY'] == 'PRI':
            ws['E{0}'.format(n)].value = 'Yes'
        if column['REFERENCED_TABLE_NAME'] is not None:
            ws['F{0}'.format(n)].value = 'Yes'
        if column['IS_NULLABLE'] == 'NO':
            ws['G{0}'.format(n)].value = 'Yes'
        ws['H{0}'.format(n)].value = column['COLUMN_COMMENT']
        ws['I{0}'.format(n)].value = column['COLUMN_TYPE']
        for row in ws['A{0}:I{0}'.format(n)]:
            for c in row:
                wrap_text = c.column in ('B', 'H')
                if c.column in ('E', 'F', 'G'):
                    horizontal = 'center'
                else:
                    horizontal = 'left'
                style_column(c, wrap_text, horizontal)

    def border_thin(cell_range):
        thin = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            top=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin'),
        )
        for row in ws[cell_range]:
            for c in row:
                c.border = thin

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    # caption
    draw_caption()

    n = 2
    for column in columns:
        draw_column(column, n)
        n += 1

    # merge cell
    n = 0
    s = 0
    table_name_old = ''
    for row in ws.rows:
        if n > 0:
            table_name = row[0].value
            if table_name != table_name_old:
                if s > 0:
                    ws.merge_cells('A{0}:A{1}'.format(s + 1, n))
                    ws.merge_cells('B{0}:B{1}'.format(s + 1, n))
                    border_thin('A{0}:A{1}'.format(s + 1, n))
                s = n
            table_name_old = table_name
        n += 1
    if n > 0:
        ws.merge_cells('A{0}:A{1}'.format(s + 1, n))
        ws.merge_cells('B{0}:B{1}'.format(s + 1, n))
        border_thin('A{0}:B{1}'.format(s + 1, n))

    wb.save(xlsx)


def main(host=DB_HOST, port=DB_PORT, user=DB_USER,
         passwd=DB_PASSWD, db=DB_NAME, charset=DB_CHARSET,
         filter=CUSTOME_FILTER, xlsx=XLSX_NAME):
    dbconnect(host=host, port=port, user=user, passwd=passwd, db=db,
              charset=charset)
    columns = fetch_columns(db=db, filter=filter)
    write_xlsx(columns, xlsx=xlsx)
    conn.close()
